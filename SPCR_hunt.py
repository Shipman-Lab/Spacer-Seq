"""Import Modules"""
from Bio import SeqIO
import os, sys
import fuzzysearch
import itertools
import xlsxwriter
from openpyxl import load_workbook

"""Line Arguments"""
run_number = sys.argv[1] #MiSeq run number for Data_Path
condition = sys.argv[2] #file name (minus .fastq) shoud be the only argument following SPCR_blast.py when running script
running_location = sys.argv[3] #either local or orchestra

"""Globals"""
if running_location == 'local':
    user_profile = os.environ ['USERPROFILE']
    Data_Path = ('%s/Dropbox (Gladstone)/eVOLVER_and_Retro_Record/MiSeq_Data/'
        + 'msSBK_%s') % (user_profile,run_number,run_number,condition) #for running locally
    # Data_Path = '%s/Dropbox/Data Analysis/MS%s/%s_Results' % (user_profile,run_number,condition) #for running locally
    Blast_Data_Path = '%s/Dropbox/Blast_Databases' % user_profile  #for running locally
elif running_location == 'orchestra_group':
    Data_Path = '/n/groups/church/Seth/MS%s_Data_Analysis/%s_Results' % (run_number,condition)  #for running on orchestra
    Blast_Data_Path = '/home/ss695/Blast_Databases' #for running on orchestra

id_list = [61, 62, 63, 64, 65, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 85, 86,
    87, 88, 89, 92, 93, 94, 95, 96]
wb_list = []

for x in id_list:
    x = load_workbook(filename = '')
    wb_list.append(load_workbook(filename = ''))
