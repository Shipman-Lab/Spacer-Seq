"""Import Modules"""
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import csv
import sys
import os
import itertools
from itertools import *

"""Line Arguments"""
resultsFolder = sys.argv[1] # MiSeq processed, trimmed results folder, output...
# from Fed_SPCRS_v2.py Eg: "C:\Users\santi.bhattaraikline\Dropbox (Gladstone)...
# \eVOLVER_and_Retro_Record\MiSeq_Data\msSBK_13\Results"

"""Globals"""
miseqRunName = os.path.split(os.path.split(resultsFolder)[0])[1]

"""Defs"""
def sampleID(trimmed_Results_folder):
	""" input the name of a Fed_SPCRS_v2 output folder (eg:
	msSBK_13_06_trimmed_Results) and return the sample ID (eg: msSBK_13_06)
	"""
	sampleID = trimmed_Results_folder.rsplit(sep="_",maxsplit=2)[0]
	return sampleID

def extractSummaryData(trimmed_Results_path,sampleID_str,entryNumber):
	""" input the path of trimmed results folder, return a list of relevant data
	from the '..._trimmed_SPCR_analysis.xlsx' and '..._oComp_Orders.xlsx' excel
	sheets
	"""
	wb_SPCR_analysis = load_workbook(filename = trimmed_Results_path + '\\'
	+ sampleID_str + '_trimmed_SPCR_analysis.xlsx', read_only=True)
	ws_SPCR_analysis = wb_SPCR_analysis.active
	wb_oComp_Orders = load_workbook(filename = trimmed_Results_path + '\\'
	+ sampleID_str + '_oComp_Orders.xlsx', read_only=True)
	ws_oComp_Orders = wb_oComp_Orders.active
	sampleData = []
	for row in ws_SPCR_analysis['B1':'B28']:
		for cell in row:
			sampleData.append(cell.value)
	for row in ws_oComp_Orders['H6':'H10']:
		for cell in row:
			sampleData.append(cell.value)
	for row in ws_oComp_Orders['B1':'B17']:
		for cell in row:
			sampleData.append(cell.value)
	for row in ws_oComp_Orders['E1':'E65']:
		for cell in row:
			sampleData.append(cell.value)

	if entryNumber == 0:
		labels = []
		for row in ws_SPCR_analysis['A1':'A28']:
			for cell in row:
				labels.append(cell.value)
		for row in ws_oComp_Orders['G6':'G10']:
			for cell in row:
				labels.append(cell.value)
		for row in ws_oComp_Orders['A1':'A17']:
			for cell in row:
				labels.append(cell.value)
		for row in ws_oComp_Orders['D1':'D65']:
			for cell in row:
				labels.append(cell.value)

		return (labels,sampleData)
	else:
		return (sampleData)

"""Run"""
# print(miseqRunName + '_summary')
workbook = xlsxwriter.Workbook(os.path.split(resultsFolder)[0] + '\\' + miseqRunName
+ '_summary.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0
entryNumber = 0

for trimmed_Results_folder in os.listdir(resultsFolder):
	trimmed_Results_path = os.path.join(resultsFolder,trimmed_Results_folder)
	sampleID_str = sampleID(trimmed_Results_folder)
	if os.path.exists(trimmed_Results_path + '\\' + sampleID_str
	+ '_trimmed_SPCR_analysis.xlsx') and os.path.exists(trimmed_Results_path
	+ '\\' + sampleID_str + '_oComp_Orders.xlsx'):
		summaryData = extractSummaryData(trimmed_Results_path,sampleID_str,
		entryNumber)
		if isinstance(summaryData, tuple):
			for value in summaryData[0]:
				worksheet.write(row,col,value)
				row += 1
			col += 1
			row = 0
			for value in summaryData[1]:
				worksheet.write(row,col,value)
				row += 1
			col += 1
		else:
			row = 0
			for value in summaryData:
				worksheet.write(row,col,value)
				row += 1
			col += 1
	entryNumber += 1

workbook.close()
	# print(trimmed_Results_folder)
	# print(sampleID_str)
	# print(os.listdir(trimmed_Results_path))
	# print()

# if range_of_possibilites['machine_lines'] == 'yes':
# 	for i in range(eval(start),eval(end)+1):
# 		with open('%s/S%s_Results/machine_lines.csv' % (Data_Path,i), 'rb') as csv_file:
# 			reader = csv.reader(csv_file)
# 			next(csv_file) #skip header row
# 			for row in reader:
# 				numbers = [int(i) for i in row]
# 				All_machine_lines.append(numbers)
#
# if range_of_possibilites['overall_excel'] == 'yes':
# 	for i in range(eval(start),eval(end)+1):
# 		sample_data = ['S%s' % i]
# 		wb1 = load_workbook(filename = '%s/S%s_Results/PAM_analysis.xlsx' % (Data_Path,i))
# 		ws = wb1.active
# 		if i == eval(start):
# 			excel_column = [ws['A1'].value]
# 			excel_column.append(ws['A2'].value)
# 		sample_data.append(ws['B1'].value)
# 		sample_data.append(ws['B2'].value)
# 		wb2 = load_workbook(filename = '%s/S%s_Results/Blast_analysis.xlsx' % (Data_Path,i))
# 		ws = wb2.active
# 		if i == eval(start):
# 			excel_column.append(ws['A2'].value)
# 			excel_column.append(ws['A3'].value)
# 		sample_data.append(ws['B2'].value)
# 		sample_data.append(ws['B3'].value)
# 		wb3 = load_workbook(filename = '%s/S%s_Results/SPCR_analysis.xlsx' % (Data_Path,i))
# 		ws = wb3.active
# 		if i == eval(start):
# 			for row in range(2,57):
# 				excel_column.append(ws['A%s' % row].value)
# 		for row in range(2,57):
# 			sample_data.append(ws['B%s' % row].value)
# 		overall_data.append(sample_data)
#
# if range_of_possibilites['plasmid_genome_lines'] == 'yes':
# 	for i in range(eval(start),eval(end)+1):
# 		with open('%s/S%s_Results/lines.csv' % (Data_Path,i), 'rb') as csv_file:
# 			reader = csv.reader(csv_file)
# 			plasmid_line_fow = map(int, next(itertools.islice(reader,2,None)))
# 			plasmid_line_rev = map(int, next(itertools.islice(reader,0,None)))
# 			BL21_genome_line_single_fow = map(int, next(itertools.islice(reader,10,None)))
# 			BL21_genome_line_single_rev = map(int, next(itertools.islice(reader,0,None)))
# 			Genome_split_single_fow = list(split_every(10000, BL21_genome_line_single_fow))
# 			Genome_split_single_rev = list(split_every(10000, BL21_genome_line_single_rev))
# 			Genome_split_line_single_fow_sum = []
# 			Genome_split_line_single_rev_sum = []
# 			for chunk in Genome_split_single_fow:
# 				Genome_split_line_single_fow_sum.append (float(sum(chunk)))
# 			for chunk in Genome_split_single_rev:
# 				Genome_split_line_single_rev_sum.append (float(sum(chunk)))
# 			line_dict['S%s_plas_f' % i] = plasmid_line_fow
# 			line_dict['S%s_plas_r' % i] = plasmid_line_rev
# 			line_dict['S%s_gen_f' % i] = Genome_split_line_single_fow_sum
# 			line_dict['S%s_gen_r' % i] = Genome_split_line_single_rev_sum
#
#
# """Output"""
# if range_of_possibilites['machine_lines'] == 'yes':
# 	with open('%s/All_machine_lines.csv' % Data_Path, 'wb') as csv_file:
# 		writer = csv.writer(csv_file)
# 		writer.writerows(All_machine_lines)
#
# if range_of_possibilites['overall_excel'] == 'yes':
# 	workbook = xlsxwriter.Workbook('%s/overall_excel.xlsx' % Data_Path)
# 	worksheet = workbook.add_worksheet()
# 	row = 1
# 	col = 0
# 	for i in range(0,len(excel_column)):
# 		worksheet.write(row,col,'%s' % excel_column[i])
# 		row += 1
# 	col += 1
# 	for sample in overall_data:
# 		row = 0
# 		for i in range(0,len(sample)):
# 			worksheet.write(row,col,'%s' % sample[i])
# 			row += 1
# 		col += 1
# 	workbook.close()
#
# if range_of_possibilites['plasmid_genome_lines'] == 'yes':
# 	wb_temp = load_workbook(filename = '%s/line_template.xlsx' % Data_Path)
# 	sheetnames = wb_temp.get_sheet_names()
# 	for i, sheet in enumerate(sheetnames):
# 		ws = wb_temp.get_sheet_by_name(sheet)
# 		for n in range(0,6068):
# 			row = n+1
# 			sample = eval(start)+i
# 			ws['D%s' % row] = line_dict['S%s_plas_f' % sample][n]
# 		for n in range(0,6068):
# 			row = n+1
# 			sample = eval(start)+i
# 			ws['F%s' % row] = line_dict['S%s_plas_r' % sample][n]
# 		for n in range(0,457):
# 			row = n+1
# 			sample = eval(start)+i
# 			ws['H%s' % row] = line_dict['S%s_gen_f' % sample][n]
# 		for n in range(0,457):
# 			row = n+1
# 			sample = eval(start)+i
# 			ws['J%s' % row] = line_dict['S%s_gen_r' % sample][n]
# 	wb_temp.save(filename = '%s/line_output%s_%s.xlsx' % (Data_Path,start,end))
#
#
#
#
#
#
#
#
